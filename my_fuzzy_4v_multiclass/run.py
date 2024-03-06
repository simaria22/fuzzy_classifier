import script_run_all_functions
import importlib
importlib.reload(script_run_all_functions)
import numpy as np
import scipy.io
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")
import time

# Record the start time
start_time = time.time()


import dataset_processing

import os
os.environ["OMP_NUM_THREADS"] = '1'
import pandas as pd
import sys

from scipy import signal

import matplotlib.pyplot as plt
import scipy.io as sio
import neurokit2 as nk
import seaborn as sns

import pandas as pd
import numpy as np
import time

# mat = 'C:/Users/mariasiouzou/Desktop/my_fuzzy_2v_multiclass/DREAMER.mat'

# # raw = sio.loadmat(mat)
file_path = 'C:/Users/mariasiouzou/Desktop/my_fuzzy_2v_multiclass/saheart.mat'
file_path1 = 'C:/Users/mariasiouzou/Desktop/my_fuzzy_2v_multiclass/saheart_gt.mat'

mat = scipy.io.loadmat(file_path)
gt = scipy.io.loadmat(file_path1)


# data = mat['data']
# data_gt = gt['data']



data_m = mat['python_mat']
data_gt1 = gt['python_gt']

# columns_to_delete = [4, 6, 7, 8, 16]
# data= np.delete(data, columns_to_delete, axis=1)

# data_gt = np.where(data_gt > 3, 1, 0)
# mask = np.isnan(data).any(axis=1) | np.isinf(data).any(axis=1)
# data_m = data[~mask]
# data_gt1= data_gt[~mask]




num_iters=1000
learning_rate=0.003
val_split=0.3
fuzzy_sets=3
n_clusters=[2,2]


accuracy, accuracy_mean, accuracy_std, accuracy_gd_mean, accuracy_gd_std, auc_dummy, auc_std, auc_gd, auc_std_gd, average_n_rules=script_run_all_functions.call_all_functions(data_m, data_gt1, num_iters, learning_rate, val_split, n_clusters, fuzzy_sets)
print("List of accuracies with dummy classification:" , accuracy)
print("Mean:", accuracy_mean, "and std:", accuracy_std, "of accuracies of dummy classification")
print("Mean:", auc_dummy, "and std:", auc_std, "of auc score of dummy classification")
# print("List of accuracies with gradient descent method:" , accuracy_gd)
print("Mean:", accuracy_gd_mean, "and std:", accuracy_gd_std, "of accuracies with gradient descent method")
print("Mean:", auc_gd, "and std:", auc_std_gd, "of auc scores with gradient descent method")

end_time = time.time()

# Calculate the total time taken
total_time = end_time - start_time
print(f"Total time taken: {total_time:.2f} seconds")

file_name = os.path.splitext(os.path.basename(file_path))[0]

# data = {
#     'Dataset' : [file_name],
#     'Accuracy Mean': [np.round(accuracy_mean, decimals=3)],
#     'Accuracy Std': [np.round(accuracy_std, decimals=3)],
#     'Accuracy GD Mean': [np.round(accuracy_gd_mean, decimals=3)],
#     'Accuracy GD Std': [np.round(accuracy_gd_std, decimals=3)],
#     'num_iters': [num_iters],
#     'learning_rate': [learning_rate],
#     'val_split': [val_split],
#     'fuzzy_sets': [fuzzy_sets],
#     'n_clusters': [n_clusters],
#     'Average number of rules' : [average_n_rules],
#     'auc' : [np.round(auc_dummy, decimals=3)],
#     'auc_std': [np.round(auc_std, decimals=3)],
#     'AUC GD' :[np.round(auc_gd, decimals=3)],
#     'AUC STD GD':[np.round(auc_std_gd, decimals=3)]
# }

# df = pd.DataFrame(data)
# output_file='experiments1.xlsx'
# # if 'cos_similarity' in file_name:
# #     output_file='results_with_cosine_similarity.xlsx'
# # else:
# #     output_file = 'output_results_experiments.xlsx'

# if os.path.isfile(output_file):
    
#     book = load_workbook(output_file)
    
#     writer = pd.ExcelWriter(output_file, engine='openpyxl')
#     writer.book = book
#     writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
#     df.to_excel(writer, sheet_name='Sheet1', index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
    
#     writer.book.save(output_file)
#     writer.book.close()
# else:
#     with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
#         df.to_excel(writer, sheet_name='Sheet1', index=False)
#         worksheet = writer.sheets['Sheet1']